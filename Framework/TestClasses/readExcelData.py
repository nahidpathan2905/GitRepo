import openpyxl

workbook = openpyxl.load_workbook("C:\\Users\mulanii\PycharmProjects\Framework\\TestData\\Nahid.xlsx")
sheet=workbook['Sheet2']     #workbook.active

#get row Size in a sheet
rowSize=sheet.max_row
print(rowSize)

#get col size
colSize=sheet.max_column
print(colSize)

#get data from sheet1
s1=sheet["A1"].value
print(s1)

print(sheet["A1"].value)

print("--------")
#alternate
s2=sheet.cell(row=1,column=1).value
print(s2)

print(sheet.cell(row=1,column=1).value)